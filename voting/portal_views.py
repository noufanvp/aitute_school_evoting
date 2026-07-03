"""
Admin Portal Views — superuser-only election management module.

Provides full CRUD for Elections, Positions, and Candidates,
including image upload handling for logos, candidate photos, and symbols.
"""
import json
import os
import uuid

from django.contrib.auth.decorators import login_required
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import transaction
from django.http import JsonResponse
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils.text import slugify
from django.views.decorators.http import require_POST

from django.contrib.auth import get_user_model
User = get_user_model()
from .models import Ballot, Candidate, Election, Position, UserProfile, Vote, KioskPresence, Student, VoterRegistration
from django.utils import timezone


# ── Helpers ──────────────────────────────────────────────────────────────────

def _superuser_required(view_fn):
    """Decorator: login + superuser check."""
    @login_required
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_superuser:
            return render(request, "portal/403.html", status=403)
        return view_fn(request, *args, **kwargs)
    _wrapped.__name__ = view_fn.__name__
    return _wrapped


def _save_upload(file_obj, folder):
    """Save an uploaded file and return the storage-assigned name.
    When Cloudinary is active, the returned name includes the version (e.g. v1782546959/...),
    which is required to correctly generate Cloudinary delivery URLs later.
    """
    if not file_obj:
        return None
    ext = os.path.splitext(file_obj.name)[1].lower() or ".bin"
    filename = f"{uuid.uuid4().hex}{ext}"
    rel_path = f"{folder}/{filename}"
    # IMPORTANT: use the return value — Cloudinary storage returns a versioned name
    saved_name = default_storage.save(rel_path, ContentFile(file_obj.read()))
    return saved_name



def _delete_media(path):
    """Delete a media-uploaded file (skip if it's a legacy static path).
    Safe to call with any path format — versioned Cloudinary paths, local paths, or empty values.
    """
    if not path:
        return
    path_str = getattr(path, "name", str(path))
    if not path_str:
        return
    # Skip static git-tracked assets — they must never be deleted
    if path_str.startswith("voting/"):
        return
    try:
        default_storage.delete(path_str)
    except Exception:
        # Silently ignore deletion errors — the DB record will still be removed
        pass


def _election_stats(election):
    """Return dict of stats for an election."""
    positions = election.positions.prefetch_related("candidates").all()
    candidate_count = sum(p.candidates.count() for p in positions)
    
    exclude_user_ids = UserProfile.objects.filter(exclude_votes=True).values_list("user_id", flat=True)
    TEST_USERNAMES = ["testuser", "micestest"]
    
    ballot_count = election.ballots.filter(
        status=Ballot.STATUS_SUBMITTED
    ).exclude(
        started_by__username__in=TEST_USERNAMES
    ).exclude(
        started_by_id__in=exclude_user_ids
    ).count()
    
    return {
        "position_count": positions.count(),
        "candidate_count": candidate_count,
        "ballot_count": ballot_count,
    }


# ── Dashboard ─────────────────────────────────────────────────────────────────

@_superuser_required
def portal_home(request):
    elections = (
        Election.objects
        .prefetch_related("positions__candidates", "ballots")
        .order_by("-created_at")
    )
    elections_data = []
    for e in elections:
        stats = _election_stats(e)
        logo_url = e.logo_url if e.logo else None
        elections_data.append({
            "obj": e,
            "stats": stats,
            "logo_url": logo_url,
            "kiosk_url": f"/vote/{e.school_slug}/" if e.school_slug else None,
        })

    context = {
        "elections_data": elections_data,
        "total_elections": len(elections_data),
        "active_count": sum(1 for ed in elections_data if ed["obj"].status == Election.STATUS_OPEN),
    }
    return render(request, "portal/home.html", context)


# ── Election CRUD ─────────────────────────────────────────────────────────────

@_superuser_required
def portal_election_create(request):
    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        school_name = request.POST.get("school_name", "").strip()
        school_slug = request.POST.get("school_slug", "").strip()
        manual_override_password = request.POST.get("manual_override_password", "").strip()
        status = request.POST.get("status", Election.STATUS_DRAFT)
        starts_at = request.POST.get("starts_at") or None
        ends_at = request.POST.get("ends_at") or None
        logo_file = request.FILES.get("logo")

        errors = {}
        if not title:
            errors["title"] = "Election title is required."
        if not school_name:
            errors["school_name"] = "School name is required."

        if errors:
            return render(request, "portal/election_form.html", {
                "errors": errors, "form": request.POST, "is_create": True,
            })

        with transaction.atomic():
            election = Election(
                title=title,
                school_name=school_name,
                school_slug=school_slug or slugify(school_name) or "school",
                status=status,
                starts_at=starts_at,
                ends_at=ends_at,
            )
            if manual_override_password:
                election.set_manual_override_password(manual_override_password)
            if logo_file:
                logo_path = _save_upload(logo_file, "elections/logos")
                election.logo = logo_path
            election.save()

        return redirect("portal-positions", election_id=election.id)

    return render(request, "portal/election_form.html", {"is_create": True, "form": {}})


@_superuser_required
def portal_election_edit(request, election_id):
    election = get_object_or_404(Election, id=election_id)

    if request.method == "POST":
        title = request.POST.get("title", "").strip()
        school_name = request.POST.get("school_name", "").strip()
        school_slug = request.POST.get("school_slug", "").strip()
        manual_override_password = request.POST.get("manual_override_password", "").strip()
        clear_manual_override_password = request.POST.get("clear_manual_override_password") == "on"
        status = request.POST.get("status", election.status)
        starts_at = request.POST.get("starts_at") or None
        ends_at = request.POST.get("ends_at") or None
        results_published = request.POST.get("results_published") == "on"
        logo_file = request.FILES.get("logo")
        remove_logo = request.POST.get("remove_logo") == "on"

        errors = {}
        if not title:
            errors["title"] = "Election title is required."
        if not school_name:
            errors["school_name"] = "School name is required."

        if errors:
            return render(request, "portal/election_form.html", {
                "errors": errors, "election": election, "is_create": False,
                "form": request.POST,
            })

        with transaction.atomic():
            election.title = title
            election.school_name = school_name
            election.school_slug = school_slug or slugify(school_name) or "school"
            election.status = status
            election.starts_at = starts_at
            election.ends_at = ends_at
            election.results_published = results_published

            if clear_manual_override_password:
                election.clear_manual_override_password()
            elif manual_override_password:
                election.set_manual_override_password(manual_override_password)

            if remove_logo and election.logo:
                _delete_media(str(election.logo))
                election.logo = None
            elif logo_file:
                old_logo = str(election.logo) if election.logo else None
                logo_path = _save_upload(logo_file, "elections/logos")
                election.logo = logo_path
                if old_logo:
                    _delete_media(old_logo)

            election.save()

        return redirect("portal-home")

    logo_url = election.logo_url if election.logo else None
    return render(request, "portal/election_form.html", {
        "election": election,
        "is_create": False,
        "logo_url": logo_url,
        "form": {
            "title": election.title,
            "school_name": election.school_name,
            "school_slug": election.school_slug,
            "status": election.status,
            "starts_at": election.starts_at.strftime("%Y-%m-%dT%H:%M") if election.starts_at else "",
            "ends_at": election.ends_at.strftime("%Y-%m-%dT%H:%M") if election.ends_at else "",
            "results_published": election.results_published,
            "has_manual_override_password": bool(election.manual_override_password_hash),
        },
    })


@_superuser_required
@require_POST
def portal_election_delete(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    if election.logo:
        _delete_media(str(election.logo))
    # Cascade deletes positions/candidates/votes via FK
    election.delete()
    return redirect("portal-home")


@_superuser_required
@require_POST
def portal_election_status(request, election_id):
    """AJAX: toggle election status."""
    election = get_object_or_404(Election, id=election_id)
    try:
        body = json.loads(request.body)
        new_status = body.get("status")
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    if new_status not in (Election.STATUS_DRAFT, Election.STATUS_OPEN, Election.STATUS_CLOSED):
        return JsonResponse({"error": "Invalid status"}, status=400)

    election.status = new_status
    election.save(update_fields=["status"])
    return JsonResponse({"status": election.status})


# ── Positions ──────────────────────────────────────────────────────────────────

@_superuser_required
def portal_positions(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    positions = election.positions.prefetch_related("candidates").order_by("order", "id")
    return render(request, "portal/positions.html", {
        "election": election,
        "positions": positions,
    })


@_superuser_required
@require_POST
def portal_position_create(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    name = request.POST.get("name", "").strip()
    icon = request.POST.get("icon", "").strip()

    if not name:
        return redirect("portal-positions", election_id=election_id)

    # Next order value
    last_order = election.positions.order_by("-order").values_list("order", flat=True).first()
    next_order = (last_order or 0) + 1

    Position.objects.create(election=election, name=name, icon=icon, order=next_order)
    return redirect("portal-positions", election_id=election_id)


@_superuser_required
def portal_position_edit(request, position_id):
    position = get_object_or_404(Position, id=position_id)
    if request.method == "POST":
        position.name = request.POST.get("name", position.name).strip()
        position.icon = request.POST.get("icon", position.icon).strip()
        position.save(update_fields=["name", "icon"])
        return redirect("portal-positions", election_id=position.election_id)

    return render(request, "portal/position_edit.html", {"position": position})


@_superuser_required
@require_POST
def portal_position_delete(request, position_id):
    position = get_object_or_404(Position, id=position_id)
    election_id = position.election_id
    # Best-effort cleanup of candidate media files — never block deletion on errors
    try:
        for candidate in position.candidates.all():
            _delete_media(candidate.photo)
            _delete_media(candidate.symbol)
    except Exception:
        pass
    position.delete()
    return redirect("portal-positions", election_id=election_id)


@_superuser_required
@require_POST
def portal_positions_reorder(request):
    """AJAX: accept ordered list of position IDs and update their order."""
    try:
        body = json.loads(request.body)
        ordered_ids = body.get("order", [])
    except (json.JSONDecodeError, ValueError):
        return JsonResponse({"error": "Invalid JSON"}, status=400)

    with transaction.atomic():
        for idx, pos_id in enumerate(ordered_ids):
            Position.objects.filter(id=pos_id).update(order=idx)

    return JsonResponse({"ok": True})


# ── Candidates ────────────────────────────────────────────────────────────────

@_superuser_required
def portal_candidate_create(request, position_id):
    position = get_object_or_404(Position, id=position_id)

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        class_name = request.POST.get("class_name", "").strip()
        motto = request.POST.get("motto", "").strip()
        is_nota = request.POST.get("is_nota") == "on"
        symbol_name = request.POST.get("symbol_name", "").strip()
        photo_file = request.FILES.get("photo")
        symbol_file = request.FILES.get("symbol")

        errors = {}
        if not name:
            errors["name"] = "Candidate name is required."

        if errors:
            return render(request, "portal/candidate_form.html", {
                "position": position,
                "errors": errors,
                "form": request.POST,
                "is_create": True,
            })

        # Next order value
        last_order = position.candidates.order_by("-order").values_list("order", flat=True).first()
        next_order = (last_order or 0) + 1

        photo_path = _save_upload(photo_file, "candidates/photos") or ""
        symbol_path = _save_upload(symbol_file, "candidates/symbols") or ""

        Candidate.objects.create(
            position=position,
            name=name,
            class_name=class_name,
            motto=motto,
            photo=photo_path,
            symbol=symbol_path,
            symbol_name=symbol_name,
            order=next_order,
            is_nota=is_nota,
        )
        return redirect("portal-positions", election_id=position.election_id)

    return render(request, "portal/candidate_form.html", {
        "position": position,
        "is_create": True,
        "form": {},
    })


@_superuser_required
def portal_candidate_edit(request, candidate_id):
    candidate = get_object_or_404(Candidate, id=candidate_id)
    position = candidate.position

    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        class_name = request.POST.get("class_name", "").strip()
        motto = request.POST.get("motto", "").strip()
        is_nota = request.POST.get("is_nota") == "on"
        symbol_name = request.POST.get("symbol_name", "").strip()
        photo_file = request.FILES.get("photo")
        symbol_file = request.FILES.get("symbol")
        remove_photo = request.POST.get("remove_photo") == "on"
        remove_symbol = request.POST.get("remove_symbol") == "on"

        errors = {}
        if not name:
            errors["name"] = "Candidate name is required."

        if errors:
            return render(request, "portal/candidate_form.html", {
                "candidate": candidate,
                "position": position,
                "errors": errors,
                "form": request.POST,
                "is_create": False,
            })

        candidate.name = name
        candidate.class_name = class_name
        candidate.motto = motto
        candidate.symbol_name = symbol_name
        candidate.is_nota = is_nota

        if remove_photo:
            _delete_media(candidate.photo)
            candidate.photo = ""
        elif photo_file:
            _delete_media(candidate.photo)
            candidate.photo = _save_upload(photo_file, "candidates/photos") or ""

        if remove_symbol:
            _delete_media(candidate.symbol)
            candidate.symbol = ""
        elif symbol_file:
            _delete_media(candidate.symbol)
            candidate.symbol = _save_upload(symbol_file, "candidates/symbols") or ""

        candidate.save()
        return redirect("portal-positions", election_id=position.election_id)

    # Build preview URLs for existing images
    def image_url(path):
        if not path:
            return None
        path_str = getattr(path, "name", str(path))
        if not path_str:
            return None
        if path_str.startswith("voting/"):
            return f"/static/{path_str}"
        return getattr(path, "url", f"/media/{path_str}")

    return render(request, "portal/candidate_form.html", {
        "candidate": candidate,
        "position": position,
        "is_create": False,
        "photo_url": image_url(candidate.photo),
        "symbol_url": image_url(candidate.symbol),
        "form": {
            "name": candidate.name,
            "class_name": candidate.class_name,
            "motto": candidate.motto,
            "symbol_name": candidate.symbol_name,
            "is_nota": candidate.is_nota,
        },
    })


@_superuser_required
@require_POST
def portal_candidate_delete(request, candidate_id):
    candidate = get_object_or_404(Candidate, id=candidate_id)
    election_id = candidate.position.election_id
    _delete_media(candidate.photo)
    _delete_media(candidate.symbol)
    candidate.delete()
    return redirect("portal-positions", election_id=election_id)


# ── Invigilators Management ──

@_superuser_required
def portal_invigilators(request):
    from django.db.models import Q
    invigilators = User.objects.filter(is_superuser=False).filter(
        Q(profile__role='invigilator') | Q(profile__isnull=True)
    ).select_related("profile").order_by("username")
    return render(request, "portal/invigilators.html", {
        "invigilators": invigilators,
    })


@_superuser_required
def portal_invigilator_create(request):
    schools = Election.objects.values_list("school_name", flat=True).distinct()
    
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        full_name = request.POST.get("full_name", "").strip()
        
        school_select = request.POST.get("school_select", "").strip()
        school_custom = request.POST.get("school_custom", "").strip()
        school_name = school_custom if school_select == "__custom__" else school_select
        
        errors = {}
        if not username:
            errors["username"] = "Username is required."
        elif User.objects.filter(username=username).exists():
            errors["username"] = "Username is already taken."
            
        if not password:
            errors["password"] = "Password is required."
        elif len(password) < 6:
            errors["password"] = "Password must be at least 6 characters."
            
        if not school_name:
            errors["school"] = "School assignment is required."
            
        if errors:
            return render(request, "portal/invigilator_form.html", {
                "errors": errors,
                "form": request.POST,
                "is_create": True,
                "schools": schools,
            })
            
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                password=password,
                is_staff=False,
                is_superuser=False,
            )
            if full_name:
                parts = full_name.split(" ", 1)
                user.first_name = parts[0]
                if len(parts) > 1:
                    user.last_name = parts[1]
                user.save()
                
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.school_name = school_name
            profile.school_slug = slugify(school_name) or "school"
            profile.exclude_votes = False
            profile.plain_password = password
            profile.role = 'invigilator'
            profile.save()
            
        return redirect("portal-invigilators")
        
    return render(request, "portal/invigilator_form.html", {
        "is_create": True,
        "form": {},
        "schools": schools,
    })


@_superuser_required
def portal_invigilator_edit(request, user_id):
    from django.db.models import Q
    user_obj = get_object_or_404(User, id=user_id, is_superuser=False)
    profile = getattr(user_obj, "profile", None)
    if not profile:
        profile, _ = UserProfile.objects.get_or_create(user=user_obj)
        profile.role = 'invigilator'
        profile.save()
        
    schools = Election.objects.values_list("school_name", flat=True).distinct()
    
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        password = request.POST.get("password", "").strip()
        
        school_select = request.POST.get("school_select", "").strip()
        school_custom = request.POST.get("school_custom", "").strip()
        school_name = school_custom if school_select == "__custom__" else school_select
        
        errors = {}
        if password and len(password) < 6:
            errors["password"] = "Password must be at least 6 characters."
            
        if not school_name:
            errors["school"] = "School assignment is required."
            
        if errors:
            return render(request, "portal/invigilator_form.html", {
                "errors": errors,
                "user_obj": user_obj,
                "profile": profile,
                "is_create": False,
                "schools": schools,
                "form": request.POST,
            })
            
        with transaction.atomic():
            if full_name:
                parts = full_name.split(" ", 1)
                user_obj.first_name = parts[0]
                user_obj.last_name = parts[1] if len(parts) > 1 else ""
            else:
                user_obj.first_name = ""
                user_obj.last_name = ""
                
            if password:
                user_obj.set_password(password)
                profile.plain_password = password
                
            user_obj.save()
            
            profile.school_name = school_name
            profile.school_slug = slugify(school_name) or "school"
            profile.exclude_votes = False
            profile.save()
            
        return redirect("portal-invigilators")
        
    full_name = user_obj.get_full_name() or user_obj.username
    return render(request, "portal/invigilator_form.html", {
        "user_obj": user_obj,
        "profile": profile,
        "is_create": False,
        "schools": schools,
        "form": {
            "username": user_obj.username,
            "full_name": full_name,
            "school_select": profile.school_name,
            "exclude_votes": profile.exclude_votes,
        },
    })


@_superuser_required
@require_POST
def portal_invigilator_delete(request, user_id):
    from django.db.models import Q
    user_obj = get_object_or_404(User, id=user_id, is_superuser=False)
    user_obj.delete()
    return redirect("portal-invigilators")


@_superuser_required
def portal_teachers(request):
    teachers = User.objects.filter(is_superuser=False, profile__role='teacher').select_related("profile").order_by("username")
    return render(request, "portal/teachers.html", {
        "teachers": teachers,
    })


@_superuser_required
def portal_teacher_create(request):
    schools = Election.objects.values_list("school_name", flat=True).distinct()
    
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        full_name = request.POST.get("full_name", "").strip()
        
        school_select = request.POST.get("school_select", "").strip()
        school_custom = request.POST.get("school_custom", "").strip()
        school_name = school_custom if school_select == "__custom__" else school_select
        
        errors = {}
        if not username:
            errors["username"] = "Username is required."
        elif User.objects.filter(username=username).exists():
            errors["username"] = "Username is already taken."
            
        if not password:
            errors["password"] = "Password is required."
        elif len(password) < 6:
            errors["password"] = "Password must be at least 6 characters."
            
        if not school_name:
            errors["school"] = "School assignment is required."
            
        if errors:
            return render(request, "portal/teacher_form.html", {
                "errors": errors,
                "form": request.POST,
                "is_create": True,
                "schools": schools,
            })
            
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                password=password,
                is_staff=False,
                is_superuser=False,
            )
            if full_name:
                parts = full_name.split(" ", 1)
                user.first_name = parts[0]
                if len(parts) > 1:
                    user.last_name = parts[1]
                user.save()
                
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.school_name = school_name
            profile.school_slug = slugify(school_name) or "school"
            profile.exclude_votes = request.POST.get("exclude_votes") == "on"
            profile.plain_password = password
            profile.role = 'teacher'
            profile.save()
            
        return redirect("portal-teachers")
        
    return render(request, "portal/teacher_form.html", {
        "is_create": True,
        "form": {},
        "schools": schools,
    })


@_superuser_required
def portal_teacher_edit(request, user_id):
    user_obj = get_object_or_404(User, id=user_id, is_superuser=False, profile__role='teacher')
    profile = getattr(user_obj, "profile", None)
    if not profile:
        profile, _ = UserProfile.objects.get_or_create(user=user_obj)
        profile.role = 'teacher'
        profile.save()
        
    schools = Election.objects.values_list("school_name", flat=True).distinct()
    
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        password = request.POST.get("password", "").strip()
        
        school_select = request.POST.get("school_select", "").strip()
        school_custom = request.POST.get("school_custom", "").strip()
        school_name = school_custom if school_select == "__custom__" else school_select
        
        errors = {}
        if password and len(password) < 6:
            errors["password"] = "Password must be at least 6 characters."
            
        if not school_name:
            errors["school"] = "School assignment is required."
            
        if errors:
            return render(request, "portal/teacher_form.html", {
                "errors": errors,
                "user_obj": user_obj,
                "profile": profile,
                "is_create": False,
                "schools": schools,
                "form": request.POST,
            })
            
        with transaction.atomic():
            if full_name:
                parts = full_name.split(" ", 1)
                user_obj.first_name = parts[0]
                user_obj.last_name = parts[1] if len(parts) > 1 else ""
            else:
                user_obj.first_name = ""
                user_obj.last_name = ""
                
            if password:
                user_obj.set_password(password)
                profile.plain_password = password
                
            user_obj.save()
            
            profile.school_name = school_name
            profile.school_slug = slugify(school_name) or "school"
            profile.exclude_votes = request.POST.get("exclude_votes") == "on"
            profile.save()
            
        return redirect("portal-teachers")
        
    full_name = user_obj.get_full_name() or user_obj.username
    return render(request, "portal/teacher_form.html", {
        "user_obj": user_obj,
        "profile": profile,
        "is_create": False,
        "schools": schools,
        "form": {
            "username": user_obj.username,
            "full_name": full_name,
            "school_select": profile.school_name,
            "exclude_votes": profile.exclude_votes,
        },
    })


@_superuser_required
@require_POST
def portal_teacher_delete(request, user_id):
    user_obj = get_object_or_404(User, id=user_id, is_superuser=False, profile__role='teacher')
    user_obj.delete()
    return redirect("portal-teachers")


@_superuser_required
def portal_kiosks(request):
    kiosks = User.objects.filter(is_superuser=False, profile__role='kiosk').select_related("profile").order_by("username")
    return render(request, "portal/kiosks.html", {
        "kiosks": kiosks,
    })


@_superuser_required
def portal_kiosk_create(request):
    schools = Election.objects.values_list("school_name", flat=True).distinct()
    
    if request.method == "POST":
        username = request.POST.get("username", "").strip()
        password = request.POST.get("password", "").strip()
        full_name = request.POST.get("full_name", "").strip()
        
        school_select = request.POST.get("school_select", "").strip()
        school_custom = request.POST.get("school_custom", "").strip()
        school_name = school_custom if school_select == "__custom__" else school_select
        
        errors = {}
        if not username:
            errors["username"] = "Kiosk Name/Username is required."
        elif User.objects.filter(username=username).exists():
            errors["username"] = "Kiosk Name/Username is already taken."
            
        if not password:
            errors["password"] = "Password/Passkey is required."
        elif len(password) < 6:
            errors["password"] = "Password must be at least 6 characters."
            
        if not school_name:
            errors["school"] = "School assignment is required."
            
        if errors:
            return render(request, "portal/kiosk_form.html", {
                "errors": errors,
                "form": request.POST,
                "is_create": True,
                "schools": schools,
            })
            
        with transaction.atomic():
            user = User.objects.create_user(
                username=username,
                password=password,
                is_staff=False,
                is_superuser=False,
            )
            if full_name:
                parts = full_name.split(" ", 1)
                user.first_name = parts[0]
                if len(parts) > 1:
                    user.last_name = parts[1]
                user.save()
                
            profile, _ = UserProfile.objects.get_or_create(user=user)
            profile.school_name = school_name
            profile.school_slug = slugify(school_name) or "school"
            profile.exclude_votes = request.POST.get("exclude_votes") == "on"
            profile.plain_password = password
            profile.role = 'kiosk'
            profile.save()
            
        return redirect("portal-kiosks")
        
    return render(request, "portal/kiosk_form.html", {
        "is_create": True,
        "form": {},
        "schools": schools,
    })


@_superuser_required
def portal_kiosk_edit(request, user_id):
    user_obj = get_object_or_404(User, id=user_id, is_superuser=False, profile__role='kiosk')
    profile = getattr(user_obj, "profile", None)
    if not profile:
        profile, _ = UserProfile.objects.get_or_create(user=user_obj)
        profile.role = 'kiosk'
        profile.save()
        
    schools = Election.objects.values_list("school_name", flat=True).distinct()
    
    if request.method == "POST":
        full_name = request.POST.get("full_name", "").strip()
        password = request.POST.get("password", "").strip()
        
        school_select = request.POST.get("school_select", "").strip()
        school_custom = request.POST.get("school_custom", "").strip()
        school_name = school_custom if school_select == "__custom__" else school_select
        
        errors = {}
        if password and len(password) < 6:
            errors["password"] = "Password must be at least 6 characters."
            
        if not school_name:
            errors["school"] = "School assignment is required."
            
        if errors:
            return render(request, "portal/kiosk_form.html", {
                "errors": errors,
                "user_obj": user_obj,
                "profile": profile,
                "is_create": False,
                "schools": schools,
                "form": request.POST,
            })
            
        with transaction.atomic():
            if full_name:
                parts = full_name.split(" ", 1)
                user_obj.first_name = parts[0]
                user_obj.last_name = parts[1] if len(parts) > 1 else ""
            else:
                user_obj.first_name = ""
                user_obj.last_name = ""
                
            if password:
                user_obj.set_password(password)
                profile.plain_password = password
                
            user_obj.save()
            
            profile.school_name = school_name
            profile.school_slug = slugify(school_name) or "school"
            profile.exclude_votes = request.POST.get("exclude_votes") == "on"
            profile.save()
            
        return redirect("portal-kiosks")
        
    full_name = user_obj.get_full_name() or user_obj.username
    return render(request, "portal/kiosk_form.html", {
        "user_obj": user_obj,
        "profile": profile,
        "is_create": False,
        "schools": schools,
        "form": {
            "username": user_obj.username,
            "full_name": full_name,
            "school_select": profile.school_name,
            "exclude_votes": profile.exclude_votes,
        },
    })


@_superuser_required
@require_POST
def portal_kiosk_delete(request, user_id):
    user_obj = get_object_or_404(User, id=user_id, is_superuser=False, profile__role='kiosk')
    user_obj.delete()
    return redirect("portal-kiosks")


@_superuser_required
def portal_active_kiosks(request):
    active_threshold = timezone.now() - timezone.timedelta(seconds=30)
    presences = KioskPresence.objects.filter(last_seen_at__gte=active_threshold).select_related("user", "election").order_by("-last_seen_at")
    
    data = []
    for p in presences:
        data.append({
            "username": p.user.username,
            "kiosk_id": p.user.username,
            "election_title": p.election.title,
            "school_name": p.election.school_name,
            "last_seen_seconds_ago": int((timezone.now() - p.last_seen_at).total_seconds()),
        })
    return JsonResponse({"active_kiosks": data})


@_superuser_required
def portal_election_turnout(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    
    # Aggregate submitted ballots by hour
    from django.db.models.functions import TruncHour
    from django.db.models import Count
    
    ballots_by_hour = (
        Ballot.objects.filter(election=election, status=Ballot.STATUS_SUBMITTED)
        .annotate(hour=TruncHour('submitted_at'))
        .values('hour')
        .annotate(count=Count('id'))
        .order_by('hour')
    )
    
    labels = []
    counts = []
    for b in ballots_by_hour:
        if b['hour']:
            # Format hour label nicely, using user's timezone if possible
            local_hour = timezone.localtime(b['hour'])
            labels.append(local_hour.strftime("%H:%M"))
            counts.append(b['count'])
            
    return JsonResponse({
        "election_title": election.title,
        "labels": labels,
        "counts": counts,
        "total_votes": sum(counts),
    })


# ── Student Roster Management ─────────────────────────────────────────────────

@_superuser_required
def portal_students(request, election_id):
    """List all students for an election with their voting status."""
    election = get_object_or_404(Election, id=election_id)
    students = election.students.all()

    # Build set of hashes that have voted
    voted_hashes = set(
        VoterRegistration.objects.filter(election=election).values_list("student_id_hash", flat=True)
    )

    import hashlib
    students_data = []
    for s in students:
        if s.student_id:
            h = hashlib.sha256(f"{election.id}:{s.student_id.lower()}".encode()).hexdigest()
            has_voted = h in voted_hashes
        else:
            has_voted = False
        students_data.append({
            "obj": s,
            "has_voted": has_voted,
        })

    total = len(students_data)
    voted_count = sum(1 for sd in students_data if sd["has_voted"])
    pending_count = total - voted_count

    return render(request, "portal/students.html", {
        "election": election,
        "students_data": students_data,
        "total": total,
        "voted_count": voted_count,
        "pending_count": pending_count,
    })


@_superuser_required
def portal_student_create(request, election_id):
    election = get_object_or_404(Election, id=election_id)
    if request.method == "POST":
        name = request.POST.get("name", "").strip()
        student_class = request.POST.get("student_class", "").strip()
        division = request.POST.get("division", "").strip()
        student_id = request.POST.get("student_id", "").strip()

        errors = {}
        if not name:
            errors["name"] = "Student name is required."
        if not student_class:
            errors["student_class"] = "Class is required."
        if student_id and Student.objects.filter(election=election, student_id__iexact=student_id).exists():
            errors["student_id"] = "A student with this ID already exists in this election."

        if errors:
            if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json' or request.POST.get('format') == 'json':
                return JsonResponse({"status": "error", "errors": errors}, status=400)
            return render(request, "portal/student_form.html", {
                "election": election, "errors": errors, "form": request.POST, "is_create": True,
            })

        student = Student.objects.create(
            election=election, name=name, student_class=student_class,
            division=division, student_id=student_id,
        )
        if request.headers.get('x-requested-with') == 'XMLHttpRequest' or request.GET.get('format') == 'json' or request.POST.get('format') == 'json':
            return JsonResponse({
                "status": "success",
                "student": {
                    "id": student.id,
                    "name": student.name,
                    "student_class": student.student_class,
                    "division": student.division,
                    "student_id": student.student_id,
                    "has_voted": False
                }
            })
        return redirect("portal-students", election_id=election_id)

    return render(request, "portal/student_form.html", {
        "election": election, "is_create": True, "form": {},
    })


@_superuser_required
def portal_student_edit(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    election = student.election
    if request.method == "POST":
        student.name = request.POST.get("name", "").strip()
        student.student_class = request.POST.get("student_class", "").strip()
        student.division = request.POST.get("division", "").strip()
        student.student_id = request.POST.get("student_id", "").strip()
        student.save()
        return redirect("portal-students", election_id=election.id)

    return render(request, "portal/student_form.html", {
        "election": election, "student": student, "is_create": False,
        "form": {
            "name": student.name,
            "student_class": student.student_class,
            "division": student.division,
            "student_id": student.student_id,
        },
    })


@_superuser_required
@require_POST
def portal_student_delete(request, student_id):
    student = get_object_or_404(Student, id=student_id)
    election_id = student.election_id
    student.delete()
    return redirect("portal-students", election_id=election_id)


@_superuser_required
def portal_student_import(request, election_id):
    """
    Excel/CSV bulk import for students.
    Expected columns (case-insensitive): Name, Class, Division, Student ID
    The header row is auto-detected.
    """
    election = get_object_or_404(Election, id=election_id)

    if request.method == "POST":
        xl_file = request.FILES.get("excel_file")
        if not xl_file:
            return render(request, "portal/student_import.html", {
                "election": election,
                "error": "Please select a file (.xlsx or .csv)",
            })

        filename_lower = xl_file.name.lower()
        rows = []
        try:
            if filename_lower.endswith(".xlsx"):
                import openpyxl
                wb = openpyxl.load_workbook(xl_file, data_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
            elif filename_lower.endswith(".csv"):
                import csv
                from io import StringIO
                file_data = xl_file.read().decode("utf-8-sig", errors="ignore")
                csv_reader = csv.reader(StringIO(file_data))
                rows = [row for row in csv_reader]
            else:
                raise ValueError("Unsupported file format. Please upload a .xlsx or .csv file.")

            if not rows:
                raise ValueError("The uploaded file is empty.")

            # Auto-detect header row
            header_row = None
            header_idx = 0
            for i, row in enumerate(rows):
                cells = [str(c).strip().lower() if c else "" for c in row]
                if "name" in cells:
                    header_row = cells
                    header_idx = i
                    break

            if header_row is None:
                raise ValueError("Could not find a header row with 'Name' column. Please use the provided template.")

            # Map column positions
            def col(names):
                for n in names:
                    if n in header_row:
                        return header_row.index(n)
                return None

            name_col = col(["name"])
            class_col = col(["class", "student_class", "class no", "classno"])
            div_col   = col(["division", "div"])
            id_col    = col(["student id", "studentid", "admission no", "admissionno", "roll no", "rollno", "id"])

            if name_col is None:
                raise ValueError("Column 'Name' is required.")
            if class_col is None:
                raise ValueError("Column 'Class' is required.")

            created = 0
            skipped = 0
            errors_list = []

            with transaction.atomic():
                for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
                    name = str(row[name_col]).strip() if row[name_col] else ""
                    student_class = str(row[class_col]).strip() if class_col is not None and row[class_col] else ""
                    division = str(row[div_col]).strip() if div_col is not None and row[div_col] else ""
                    student_id = str(row[id_col]).strip() if id_col is not None and row[id_col] else ""

                    # Clean up float-style IDs like "1042.0" → "1042"
                    if student_id.endswith(".0") and student_id[:-2].isdigit():
                        student_id = student_id[:-2]
                    if student_class.endswith(".0") and student_class[:-2].isdigit():
                        student_class = student_class[:-2]

                    if not name or not student_class:
                        skipped += 1
                        continue

                    try:
                        Student.objects.get_or_create(
                            election=election,
                            student_id=student_id,
                            defaults={
                                "name": name,
                                "student_class": student_class,
                                "division": division,
                            },
                        ) if student_id else Student.objects.create(
                            election=election,
                            name=name,
                            student_class=student_class,
                            division=division,
                            student_id=student_id,
                        )
                        created += 1
                    except Exception as e:
                        errors_list.append(f"Row {row_num}: {e}")
                        skipped += 1

            return render(request, "portal/student_import.html", {
                "election": election,
                "success": True,
                "created": created,
                "skipped": skipped,
                "errors_list": errors_list[:20],
            })

        except Exception as e:
            return render(request, "portal/student_import.html", {
                "election": election,
                "error": str(e),
            })

    return render(request, "portal/student_import.html", {"election": election})


@_superuser_required
def portal_student_import_template(request, election_id):
    """Download an .xlsx or .csv template file for student bulk import."""
    election = get_object_or_404(Election, id=election_id)
    file_format = request.GET.get("format", "xlsx").lower()

    if file_format == "csv":
        import csv
        
        response = HttpResponse(content_type="text/csv; charset=utf-8")
        filename = f"student_import_template_{election.school_slug or election.id}.csv"
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        
        writer = csv.writer(response)
        writer.writerow(["Name", "Class", "Division", "Student ID"])
        return response

    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Students"
    ws.append(["Name", "Class", "Division", "Student ID"])

    # Optional readability tweaks
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 18

    from io import BytesIO

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"student_import_template_{election.school_slug or election.id}.xlsx"
    response = HttpResponse(
        output.getvalue(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


@_superuser_required
def portal_students_api(request, election_id):
    """JSON API — returns student list with voting status for invigilator dashboard."""
    import hashlib
    election = get_object_or_404(Election, id=election_id)
    students = election.students.all()

    voted_hashes = set(
        VoterRegistration.objects.filter(election=election).values_list("student_id_hash", flat=True)
    )

    data = []
    for s in students:
        if s.student_id:
            h = hashlib.sha256(f"{election.id}:{s.student_id.lower()}".encode()).hexdigest()
            has_voted = h in voted_hashes
        else:
            has_voted = False
        data.append({
            "id": s.id,
            "name": s.name,
            "student_class": s.student_class,
            "division": s.division,
            "student_id": s.student_id,
            "has_voted": has_voted,
        })

    return JsonResponse({"students": data, "election_id": election_id})
