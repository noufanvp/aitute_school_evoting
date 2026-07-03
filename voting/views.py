import json
import hashlib
import csv
from io import StringIO
from io import BytesIO

from django.contrib.auth import get_user_model
from django.contrib.auth.decorators import login_required
from django.db import transaction
from django.db.models import Count
from django.http import HttpResponse
from django.http import HttpResponseBadRequest
from django.http import HttpResponseForbidden
from django.http import JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.utils import timezone
from django.views.decorators.http import require_GET
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

from .models import Ballot
from .models import Candidate
from .models import Election
from .models import Position
from .models import Student
from .models import UserProfile
from .models import Vote
from .models import VoterRegistration
from .models import KioskPresence
from .models import KioskSession
from .rate_limit import throttle_request

User = get_user_model()


def _active_election_or_none(school_slug=None):
	elections = Election.objects.filter(status=Election.STATUS_OPEN).order_by("-created_at")
	if school_slug:
		elections = elections.filter(school_slug=school_slug)
	return elections.first()


def _student_for_hash(election, student_hash):
	"""Resolve student details from a stored hash using election-bound student IDs."""
	if not student_hash:
		return None

	for student in Student.objects.filter(election=election):
		val = student.student_id.lower() if student.student_id else f"student_id__{student.id}"
		computed = hashlib.sha256(f"{election.id}:{val}".encode("utf-8")).hexdigest()
		if computed == student_hash:
			return {
				"name": student.name,
				"student_class": student.student_class,
				"division": student.division,
				"student_id": student.student_id or "",
			}
	return None


def _serialize_election(election):

	def _resolve_field_url(image_field):
		"""Return the correct URL for an image field.
		- Static assets (voting/...) → /static/voting/...
		- Cloudinary uploads → full https://res.cloudinary.com/... URL
		- Legacy local uploads → /media/... path
		"""
		if not image_field:
			return ""
		name = image_field.name or ""
		if name.startswith("voting/"):
			return f"/static/{name}"
		try:
			return image_field.url  # Returns full Cloudinary URL or local /media/ URL
		except (ValueError, AttributeError):
			return ""

	positions = []
	for position in election.positions.prefetch_related("candidates").all().order_by("order", "id"):
		positions.append(
			{
				"id": position.id,
				"position": position.name,
				"icon": position.icon,
				"candidates": [
					{
						"id": candidate.id,
						"name": candidate.name,
						"class": candidate.class_name,
						"motto": candidate.motto,
						"photo": _resolve_field_url(candidate.photo),
						"symbol": _resolve_field_url(candidate.symbol),
						"symbol_name": candidate.symbol_name,
					}
					for candidate in position.candidates.all().order_by("order", "id")
				],
			}
		)

	return {
		"id": election.id,
		"title": election.title,
		"school_name": election.school_name,
		"school_slug": election.school_slug,
		"logo_url": election.logo_url,
		"kiosk_timeout": election.kiosk_timeout,
		"positions": positions,
	}


def _cell(row, index):
	if index is None or index >= len(row):
		return ""
	value = row[index]
	return str(value).strip() if value is not None else ""


def _import_students_from_upload(election, upload):
	"""
	Import students from .xlsx/.csv using the same format as the admin portal.
	Expected columns: Name, Class, Division, Student ID.
	"""
	if not upload:
		raise ValueError("Please select a file (.xlsx or .csv).")

	filename_lower = upload.name.lower()
	if filename_lower.endswith(".xlsx"):
		import openpyxl
		wb = openpyxl.load_workbook(upload, data_only=True)
		ws = wb.active
		rows = list(ws.iter_rows(values_only=True))
	elif filename_lower.endswith(".csv"):
		file_data = upload.read().decode("utf-8-sig", errors="ignore")
		rows = [row for row in csv.reader(StringIO(file_data))]
	else:
		raise ValueError("Unsupported file format. Please upload a .xlsx or .csv file.")

	if not rows:
		raise ValueError("The uploaded file is empty.")

	header_row = None
	header_idx = 0
	for i, row in enumerate(rows):
		cells = [str(c).strip().lower() if c else "" for c in row]
		if "name" in cells:
			header_row = cells
			header_idx = i
			break

	if header_row is None:
		raise ValueError("Could not find a header row with 'Name' column.")

	def col(names):
		for name in names:
			if name in header_row:
				return header_row.index(name)
		return None

	name_col = col(["name"])
	class_col = col(["class", "student_class", "class no", "classno"])
	div_col = col(["division", "div"])
	id_col = col(["student id", "studentid", "admission no", "admissionno", "roll no", "rollno", "id"])

	if name_col is None:
		raise ValueError("Column 'Name' is required.")
	if class_col is None:
		raise ValueError("Column 'Class' is required.")

	created = 0
	skipped = 0
	errors = []

	with transaction.atomic():
		for row_num, row in enumerate(rows[header_idx + 1:], start=header_idx + 2):
			name = _cell(row, name_col)
			student_class = _cell(row, class_col)
			division = _cell(row, div_col)
			student_id = _cell(row, id_col)

			if student_id.endswith(".0") and student_id[:-2].isdigit():
				student_id = student_id[:-2]
			if student_class.endswith(".0") and student_class[:-2].isdigit():
				student_class = student_class[:-2]

			if not name or not student_class:
				skipped += 1
				continue

			try:
				if student_id:
					student, was_created = Student.objects.get_or_create(
						election=election,
						student_id=student_id,
						defaults={
							"name": name,
							"student_class": student_class,
							"division": division,
						},
					)
					if was_created:
						created += 1
					else:
						skipped += 1
				else:
					Student.objects.create(
						election=election,
						name=name,
						student_class=student_class,
						division=division,
						student_id=student_id,
					)
					created += 1
			except Exception as exc:
				errors.append(f"Row {row_num}: {exc}")
				skipped += 1

	return {
		"created": created,
		"skipped": skipped,
		"errors": errors[:20],
	}


def _teacher_student_election(request):
	profile = getattr(request.user, 'profile', None)
	if not request.user.is_superuser:
		if not profile or profile.is_locked or profile.role != 'teacher':
			return None, JsonResponse({"detail": "Access denied. Only teacher accounts can manage students."}, status=403)

	school_slug = getattr(profile, 'school_slug', '') if profile else ''
	if request.user.is_superuser:
		school_slug = request.POST.get("school_slug", "").strip() or request.GET.get("school_slug", "").strip() or school_slug or None

	election = _active_election_or_none(school_slug=school_slug or None)
	if not election:
		return None, JsonResponse({"detail": "No active election found for your school."}, status=400)
	return election, None


def _student_import_template_response(election, file_format):
	headers = ["Name", "Class", "Division", "Student ID"]
	suffix = election.school_slug or election.id

	if file_format == "csv":
		response = HttpResponse(content_type="text/csv; charset=utf-8")
		response["Content-Disposition"] = f'attachment; filename="student_import_template_{suffix}.csv"'
		writer = csv.writer(response)
		writer.writerow(headers)
		return response

	import openpyxl

	wb = openpyxl.Workbook()
	ws = wb.active
	ws.title = "Students"
	ws.append(headers)
	ws.column_dimensions["A"].width = 28
	ws.column_dimensions["B"].width = 12
	ws.column_dimensions["C"].width = 12
	ws.column_dimensions["D"].width = 18

	output = BytesIO()
	wb.save(output)
	output.seek(0)

	response = HttpResponse(
		output.getvalue(),
		content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
	)
	response["Content-Disposition"] = f'attachment; filename="student_import_template_{suffix}.xlsx"'
	return response


@login_required
@require_GET
def kiosk_page(request, school_slug=None):
	profile = getattr(request.user, "profile", None)
	if not request.user.is_superuser:
		if profile and profile.role in ["invigilator", "teacher"]:
			return redirect("invigilator-dashboard")
		if not profile or profile.role != "kiosk":
			return HttpResponseForbidden("Access denied. Kiosk access is restricted to kiosk devices.")
		if profile.school_slug:
			if school_slug != profile.school_slug:
				return redirect("kiosk-school", school_slug=profile.school_slug)

	election = _active_election_or_none(school_slug=school_slug)
	is_test_kiosk = request.user.is_superuser or request.user.username in ["testuser", "micestest"] or (
		profile and profile.exclude_votes
	)

	response_cookie_id = None
	cookie_device_id = request.COOKIES.get('kiosk_device_id', '')
	if not cookie_device_id:
		import uuid
		cookie_device_id = str(uuid.uuid4())
		response_cookie_id = cookie_device_id

	show_takeover_prompt = False
	if profile and not request.user.is_superuser:
		# Check if another device has been active recently (within 15s)
		active_threshold = timezone.now() - timezone.timedelta(seconds=15)
		presence_exists = KioskPresence.objects.filter(
			election=election,
			user=request.user,
			last_seen_at__gte=active_threshold
		).exists() if election else False

		if profile.current_device_id and profile.current_device_id != cookie_device_id and presence_exists:
			# Another laptop is actively using this kiosk user
			show_takeover_prompt = True
		else:
			# No active session or it's the same device, register this device
			profile.current_device_id = cookie_device_id
			profile.save(update_fields=["current_device_id"])

	response = render(request, "voting/index.html", {
		"election": election,
		"school_slug": school_slug or "",
		"is_test_kiosk": is_test_kiosk,
		"show_takeover_prompt": show_takeover_prompt,
	})

	if response_cookie_id:
		response.set_cookie('kiosk_device_id', response_cookie_id, max_age=31536000, samesite='Lax')

	return response


@login_required
@require_POST
def api_kiosk_takeover(request):
	profile = getattr(request.user, "profile", None)
	if not profile or profile.role != "kiosk":
		return JsonResponse({"detail": "Access denied."}, status=403)

	cookie_device_id = request.COOKIES.get('kiosk_device_id', '')
	if not cookie_device_id:
		return JsonResponse({"detail": "Missing device identifier."}, status=400)

	profile.current_device_id = cookie_device_id
	profile.save(update_fields=["current_device_id"])

	# Also delete the old presence to reset timer
	KioskPresence.objects.filter(user=request.user).delete()

	return JsonResponse({"status": "ok"})


@login_required
@require_GET
def api_current_election(request):
	school_slug = request.GET.get("school") or None
	if not request.user.is_superuser:
		profile = getattr(request.user, "profile", None)
		if profile and profile.school_slug:
			school_slug = profile.school_slug

	election = _active_election_or_none(school_slug=school_slug)
	if not election:
		return JsonResponse({"detail": "No active election."}, status=404)
	return JsonResponse(_serialize_election(election), status=200)


def _check_kiosk_device(request):
	profile = getattr(request.user, 'profile', None)
	if profile and profile.role == 'kiosk':
		cookie_device_id = request.COOKIES.get('kiosk_device_id', '')
		if profile.current_device_id and profile.current_device_id != cookie_device_id:
			return JsonResponse({"status": "logged_out", "detail": "This kiosk has been opened on another device."}, status=401)
	return None


@login_required
@require_POST
def api_start_session(request):
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	if throttle_request(request, "start_session", limit=20, window_seconds=60):
		return JsonResponse({"detail": "Too many start-session attempts."}, status=429)

	# Block locked invigilators
	profile = getattr(request.user, "profile", None)
	if profile and profile.is_locked:
		return JsonResponse(
			{"detail": "Your account has been locked by the administrator. No new ballots can be started."},
			status=423,
		)

	# Parse body (optional — may be empty JSON or contain school_slug)
	try:
		payload = json.loads(request.body.decode("utf-8")) if request.body else {}
	except json.JSONDecodeError:
		payload = {}

	school_slug = payload.get("school_slug") or None
	if not request.user.is_superuser:
		profile = getattr(request.user, "profile", None)
		if profile and profile.school_slug:
			school_slug = profile.school_slug

	election = _active_election_or_none(school_slug=school_slug)
	if not election:
		return JsonResponse({"detail": "Election is not open."}, status=400)
	if not election.is_active():
		return JsonResponse({"detail": f"Election is scheduled but not active yet. Starts at: {election.starts_at}"}, status=400)

	# Turnout safeguard: student_id validation and hashing
	is_test_session = request.user.is_superuser or request.user.username in ["testuser", "micestest"] or (
		profile and profile.exclude_votes
	)
	student_id = payload.get("student_id", "").strip().lower()
	manual_override_password = payload.get("manual_override_password", "").strip()
	kiosk_session_id = payload.get("kiosk_session_id")

	# ── Invigilator-activated path ─────────────────────────────────
	# If a kiosk_session_id is provided, the invigilator already validated
	# the student. Skip the student ID modal validation and link to KioskSession.
	kiosk_session = None
	if kiosk_session_id:
		try:
			kiosk_session = KioskSession.objects.get(id=kiosk_session_id, status=KioskSession.STATUS_ACTIVE)
		except KioskSession.DoesNotExist:
			return JsonResponse({"detail": "Session expired or invalid. Please ask the invigilator to activate again."}, status=400)
		if kiosk_session.election_id != election.id:
			return JsonResponse({"detail": "Session does not match the active school election."}, status=400)

	# ── Manual / legacy path ───────────────────────────────────────
	elif not is_test_session and not manual_override_password and not student_id:
		if election.manual_override_password_hash:
			return JsonResponse({"detail": "Manual override password is required to start a voter session."}, status=400)
		return JsonResponse({"detail": "Student ID is required to start a voter session."}, status=400)

	student_hash = None
	student_obj = None
	if kiosk_session:
		# Use hash from the KioskSession (already computed by invigilator panel)
		student_hash = kiosk_session.student_id_hash
		is_test_session = kiosk_session.activated_by and (
			kiosk_session.activated_by.is_superuser or
			getattr(getattr(kiosk_session.activated_by, "profile", None), "exclude_votes", False)
		)
		student_obj = _student_for_hash(election, student_hash)
	elif manual_override_password:
		if not election.manual_override_password_hash:
			return JsonResponse({"detail": "Manual override is not configured for this school."}, status=400)
		if not election.check_manual_override_password(manual_override_password):
			return JsonResponse({"detail": "Invalid manual override password."}, status=403)
		if student_id:
			import hashlib
			student_hash = hashlib.sha256(f"{election.id}:{student_id}".encode("utf-8")).hexdigest()
			student_obj = _student_for_hash(election, student_hash)
	elif student_id:
		import hashlib
		student_hash = hashlib.sha256(f"{election.id}:{student_id}".encode("utf-8")).hexdigest()
		if not is_test_session:
			if VoterRegistration.objects.filter(election=election, student_id_hash=student_hash).exists():
				return JsonResponse({"detail": "This student has already voted in this election."}, status=400)
		student_obj = _student_for_hash(election, student_hash)

	if not student_obj and is_test_session:
		student_obj = {
			"name": student_id.upper() if student_id else "TEST STUDENT",
			"student_class": "TEST",
			"division": "SESSION",
			"student_id": student_id or "TEST",
		}

	with transaction.atomic():

		ballot = Ballot.objects.create(
			election=election,
			started_by=request.user,
			session_token=Ballot.generate_session_token(),
			receipt_token=Ballot.generate_receipt_token(),
		)

		# Link ballot to KioskSession if applicable
		if kiosk_session:
			kiosk_session.ballot = ballot
			kiosk_session.save(update_fields=["ballot"])
			request.session["active_kiosk_session_id"] = kiosk_session.id
		elif student_hash:
			request.session["active_student_id_hash"] = student_hash

	request.session["active_ballot_id"] = ballot.id

	return JsonResponse(
		{
			"ballot_id": ballot.id,
			"session_token": ballot.session_token,
			"election": _serialize_election(election),
			"student": student_obj,
		},
		status=201,
	)


@login_required
@require_POST
def api_save_selection(request):
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	if throttle_request(request, "save_selection", limit=120, window_seconds=60):
		return JsonResponse({"detail": "Too many vote updates."}, status=429)

	try:
		payload = json.loads(request.body.decode("utf-8"))
	except json.JSONDecodeError:
		return HttpResponseBadRequest("Invalid JSON payload.")

	ballot_id = payload.get("ballot_id")
	position_id = payload.get("position_id")
	candidate_id = payload.get("candidate_id")

	if not all([ballot_id, position_id, candidate_id]):
		return JsonResponse({"detail": "ballot_id, position_id and candidate_id are required."}, status=400)

	try:
		ballot = Ballot.objects.select_related("election").get(id=ballot_id)
	except Ballot.DoesNotExist:
		return JsonResponse({"detail": "Ballot not found."}, status=404)

	if request.session.get("active_ballot_id") != ballot.id:
		return JsonResponse({"detail": "This ballot session is not active on kiosk."}, status=403)

	ballot_token = request.headers.get("X-Ballot-Token")
	if not ballot_token or ballot_token != ballot.session_token:
		return JsonResponse({"detail": "Invalid or expired ballot token."}, status=403)

	if ballot.status != Ballot.STATUS_STARTED:
		return JsonResponse({"detail": "Ballot is already submitted or cancelled."}, status=400)

	if not ballot.election.is_active():
		return JsonResponse({"detail": "Election is no longer active."}, status=400)

	try:
		position = Position.objects.get(id=position_id, election=ballot.election)
	except Position.DoesNotExist:
		return JsonResponse({"detail": "Position does not belong to active election."}, status=400)

	try:
		candidate = Candidate.objects.get(id=candidate_id, position=position)
	except Candidate.DoesNotExist:
		return JsonResponse({"detail": "Candidate does not belong to selected position."}, status=400)

	Vote.objects.update_or_create(
		ballot=ballot,
		position=position,
		defaults={"candidate": candidate},
	)

	return JsonResponse({"detail": "Selection saved."}, status=200)


@login_required
@require_POST
def api_submit_ballot(request):
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	if throttle_request(request, "submit_ballot", limit=30, window_seconds=60):
		return JsonResponse({"detail": "Too many submit attempts."}, status=429)

	# Block locked invigilators
	profile = getattr(request.user, "profile", None)
	if profile and profile.is_locked:
		return JsonResponse(
			{"detail": "Your account has been locked by the administrator. Ballot submission is disabled."},
			status=423,
		)

	try:
		payload = json.loads(request.body.decode("utf-8"))
	except json.JSONDecodeError:
		return HttpResponseBadRequest("Invalid JSON payload.")

	ballot_id = payload.get("ballot_id")
	if not ballot_id:
		return JsonResponse({"detail": "ballot_id is required."}, status=400)

	try:
		ballot = Ballot.objects.select_related("election").get(id=ballot_id)
	except Ballot.DoesNotExist:
		return JsonResponse({"detail": "Ballot not found."}, status=404)

	if request.session.get("active_ballot_id") != ballot.id:
		return JsonResponse({"detail": "This ballot session is not active on kiosk."}, status=403)

	ballot_token = request.headers.get("X-Ballot-Token")
	if not ballot_token or ballot_token != ballot.session_token:
		return JsonResponse({"detail": "Invalid or expired ballot token."}, status=403)

	if ballot.status != Ballot.STATUS_STARTED:
		return JsonResponse({"detail": "Ballot was already submitted."}, status=409)

	if not ballot.election.is_active():
		return JsonResponse({"detail": "Election is no longer active."}, status=400)

	position_count = ballot.election.positions.count()
	vote_count = ballot.votes.count()
	if vote_count != position_count:
		return JsonResponse(
			{"detail": "All positions must be voted before submit.", "expected": position_count, "current": vote_count},
			status=400,
		)

	with transaction.atomic():
		ballot = Ballot.objects.select_for_update().get(id=ballot.id)
		if ballot.status != Ballot.STATUS_STARTED:
			return JsonResponse({"detail": "Ballot was already submitted."}, status=409)

		ballot.status = Ballot.STATUS_SUBMITTED
		ballot.submitted_at = timezone.now()
		ballot.save(update_fields=["status", "submitted_at"])

	# Mark linked KioskSession as done so invigilator polling detects completion.
	# Also register the voter HERE (not at session-start) so that timeout/cancellation
	# does NOT mark the student as already voted.
	is_test = ballot.started_by and (
		ballot.started_by.is_superuser or
		getattr(getattr(ballot.started_by, 'profile', None), 'exclude_votes', False)
	)

	# Try ballot-linked KioskSession first
	linked_session = None
	try:
		linked_session = KioskSession.objects.get(ballot=ballot)
	except KioskSession.DoesNotExist:
		pass

	# Fallback: session stored by session-start flow
	if not linked_session:
		kiosk_session_id = request.session.get("active_kiosk_session_id")
		if kiosk_session_id:
			try:
				linked_session = KioskSession.objects.get(id=kiosk_session_id)
			except KioskSession.DoesNotExist:
				pass

	if linked_session:
		linked_session.status = KioskSession.STATUS_DONE
		linked_session.save(update_fields=["status"])
		# Register voter using hash stored on the KioskSession
		if not is_test and linked_session.student_id_hash:
			VoterRegistration.objects.get_or_create(
				election=linked_session.election,
				student_id_hash=linked_session.student_id_hash,
			)
	else:
		# Fallback: register voter using student_id_hash stored in session for manual flow
		student_id_hash = request.session.get("active_student_id_hash")
		if student_id_hash and not is_test:
			VoterRegistration.objects.get_or_create(
				election=ballot.election,
				student_id_hash=student_id_hash,
			)

	request.session.pop("active_kiosk_session_id", None)
	request.session.pop("active_student_id_hash", None)
	request.session.pop("active_ballot_id", None)
	return JsonResponse({"receipt_token": ballot.receipt_token}, status=200)


@require_GET
def results_page(request, election_id):
	election = get_object_or_404(Election, id=election_id)

	# Access control: unpublished results are staff/superuser only
	if not election.results_published:
		if not request.user.is_authenticated or not (request.user.is_staff or request.user.is_superuser):
			return HttpResponseForbidden(
				"<h2 style='font-family:sans-serif;text-align:center;margin-top:80px'>"
				"🔒 Results have not been published yet.</h2>"
			)

	# Test accounts — their votes are excluded from official results
	TEST_USERNAMES = ["testuser", "micestest"]
	exclude_user_ids = UserProfile.objects.filter(exclude_votes=True).values_list("user_id", flat=True)

	# Count only ballots that have actual votes in this election (excludes stale/orphaned ballots)
	total_submitted = (
		Vote.objects.filter(position__election=election, ballot__status=Ballot.STATUS_SUBMITTED)
		.exclude(ballot__started_by__username__in=TEST_USERNAMES)
		.exclude(ballot__started_by_id__in=exclude_user_ids)
		.values("ballot_id")
		.distinct()
		.count()
	)

	# Optimize queries: fetch all vote counts for all positions/candidates in a single query
	all_vote_counts = (
		Vote.objects.filter(position__election=election, ballot__status=Ballot.STATUS_SUBMITTED)
		.exclude(ballot__started_by__username__in=TEST_USERNAMES)
		.exclude(ballot__started_by_id__in=exclude_user_ids)
		.values("position_id", "candidate_id")
		.annotate(count=Count("id"))
	)

	# Build a nested dictionary: position_id -> candidate_id -> count
	counts_map = {}
	for vc in all_vote_counts:
		pos_id = vc["position_id"]
		cand_id = vc["candidate_id"]
		count = vc["count"]
		if pos_id not in counts_map:
			counts_map[pos_id] = {}
		counts_map[pos_id][cand_id] = count

	positions_data = []
	for position in election.positions.prefetch_related("candidates").order_by("order", "id"):
		pos_counts = counts_map.get(position.id, {})
		total_votes = sum(pos_counts.values())

		candidates_data = []
		for candidate in position.candidates.order_by("order", "id"):
			votes = pos_counts.get(candidate.id, 0)
			pct = round((votes / total_votes * 100) if total_votes else 0, 1)
			candidates_data.append({
				"id": candidate.id,
				"name": candidate.name,
				"photo": candidate.photo,
				"symbol": candidate.symbol,
				"class_name": candidate.class_name,
				"votes": votes,
				"pct": pct,
			})

		# Sort descending by votes to find winner
		candidates_data.sort(key=lambda c: c["votes"], reverse=True)
		winner_votes = candidates_data[0]["votes"] if candidates_data else 0
		for c in candidates_data:
			c["is_winner"] = (c["votes"] == winner_votes and winner_votes > 0)

		positions_data.append({
			"name": position.name,
			"icon": position.icon,
			"total_votes": total_votes,
			"candidates": candidates_data,
		})

	context = {
		"election": election,
		"positions": positions_data,
		"total_submitted": total_submitted,
		"is_superuser": request.user.is_authenticated and request.user.is_superuser,
	}
	return render(request, "voting/results.html", context)


@login_required
@require_POST
def api_publish_results(request, election_id):
	if not request.user.is_superuser:
		return JsonResponse({"detail": "Superuser access required."}, status=403)

	election = get_object_or_404(Election, id=election_id)
	election.results_published = not election.results_published
	election.save(update_fields=["results_published"])
	return JsonResponse({"published": election.results_published}, status=200)


@login_required
@require_GET
def api_list_invigilators(request):
	"""Return list of all non-superuser users with their lock status (superuser only)."""
	if not request.user.is_superuser:
		return JsonResponse({"detail": "Superuser access required."}, status=403)

	users = (
		User.objects
		.filter(is_superuser=False)
		.select_related("profile")
		.order_by("username")
	)

	data = []
	for u in users:
		profile = getattr(u, "profile", None)
		if profile is None:
			profile, _ = UserProfile.objects.get_or_create(user=u)
		data.append({
			"id": u.id,
			"username": u.username,
			"full_name": u.get_full_name() or u.username,
			"is_staff": u.is_staff,
			"is_active": u.is_active,
			"is_locked": profile.is_locked,
			"locked_reason": profile.locked_reason,
			"locked_at": profile.locked_at.isoformat() if profile.locked_at else None,
			"school_name": profile.school_name,
			"school_slug": profile.school_slug,
			"exclude_votes": profile.exclude_votes,
		})

	return JsonResponse({"users": data}, status=200)


@login_required
@require_POST
def api_toggle_user_lock(request, user_id):
	"""Lock or unlock a specific invigilator user (superuser only)."""
	if not request.user.is_superuser:
		return JsonResponse({"detail": "Superuser access required."}, status=403)

	target_user = get_object_or_404(User, id=user_id)

	if target_user.is_superuser:
		return JsonResponse({"detail": "Cannot lock a superuser account."}, status=400)

	try:
		payload = json.loads(request.body.decode("utf-8"))
	except json.JSONDecodeError:
		payload = {}

	reason = payload.get("reason", "Locked by administrator for result publishing")

	profile, _ = UserProfile.objects.get_or_create(user=target_user)
	profile.is_locked = not profile.is_locked
	if profile.is_locked:
		profile.locked_at = timezone.now()
		profile.locked_reason = reason
	else:
		profile.locked_at = None
		profile.locked_reason = ""
	profile.save()

	return JsonResponse({
		"user_id": target_user.id,
		"username": target_user.username,
		"is_locked": profile.is_locked,
	}, status=200)


def verify_ballot_page(request):
	token = request.GET.get("token", "").strip()
	ballot = None
	searched = False
	error_message = None

	if token:
		searched = True
		try:
			ballot = Ballot.objects.select_related("election").get(
				receipt_token=token,
				status=Ballot.STATUS_SUBMITTED
			)
		except Ballot.DoesNotExist:
			error_message = "Invalid or unsubmitted ballot receipt token."

	return render(
		request,
		"voting/verify.html",
		{
			"token": token,
			"ballot": ballot,
			"searched": searched,
			"error_message": error_message,
		},
	)


@csrf_exempt
@login_required
@require_POST
def api_kiosk_ping(request):
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	try:
		payload = json.loads(request.body.decode("utf-8")) if request.body else {}
	except Exception:
		payload = {}
	school_slug = payload.get("school_slug")
	kiosk_id = payload.get("kiosk_id", "")
	election = _active_election_or_none(school_slug=school_slug)
	if not election:
		return JsonResponse({"detail": "No active election found for ping."}, status=400)

	KioskPresence.objects.update_or_create(
		election=election,
		user=request.user,
		defaults={"last_seen_at": timezone.now()}
	)
	return JsonResponse({"status": "ok", "kiosk_id": kiosk_id})


@login_required
def api_kiosk_session_check(request):
	"""Kiosk polls this every 2 seconds to find out if an invigilator has assigned it a voter session."""
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	kiosk_id = request.GET.get("kiosk_id", "").strip()
	school_slug = request.GET.get("school_slug", "").strip()

	if not kiosk_id:
		return JsonResponse({"status": "idle"})

	# Expire old sessions
	now = timezone.now()
	pending_qs = KioskSession.objects.filter(
		kiosk_id=kiosk_id,
		status=KioskSession.STATUS_PENDING,
	)
	if school_slug:
		pending_qs = pending_qs.filter(election__school_slug=school_slug)

	pending_qs.filter(expires_at__lt=now).update(status=KioskSession.STATUS_EXPIRED)

	# Look for a pending session for this kiosk
	try:
		session_qs = KioskSession.objects.filter(
			kiosk_id=kiosk_id,
			status=KioskSession.STATUS_PENDING,
		)
		if school_slug:
			session_qs = session_qs.filter(election__school_slug=school_slug)

		session = session_qs.latest("created_at")
		# Mark it as active
		session.status = KioskSession.STATUS_ACTIVE
		session.activated_at = now
		session.save(update_fields=["status", "activated_at"])

		election = session.election
		student = _student_for_hash(election, session.student_id_hash)
		return JsonResponse({
			"status": "session_ready",
			"session_id": session.id,
			"election": _serialize_election(election),
			"student": student,
			"countdown": max(0, int((session.expires_at - now).total_seconds())),
		})
	except KioskSession.DoesNotExist:
		return JsonResponse({"status": "idle"})


@login_required
@require_POST
def api_invigilator_activate_kiosk(request):
	"""Invigilator activates a specific kiosk for a specific student."""
	profile = getattr(request.user, 'profile', None)
	if not request.user.is_superuser:
		if not profile or profile.is_locked or profile.role != 'invigilator':
			return JsonResponse({"detail": "Access denied. Only invigilator accounts can activate kiosks."}, status=403)

	try:
		payload = json.loads(request.body.decode("utf-8"))
	except Exception:
		return JsonResponse({"detail": "Invalid JSON."}, status=400)

	student_id = payload.get("student_id", "").strip().lower()
	kiosk_id = payload.get("kiosk_id", "").strip()
	school_slug = payload.get("school_slug", "").strip()

	if not student_id or not kiosk_id:
		return JsonResponse({"detail": "student_id and kiosk_id are required."}, status=400)

	# Find active election
	election = _active_election_or_none(school_slug=school_slug or None)
	if not election:
		return JsonResponse({"detail": "No active election found."}, status=400)

	# Hash the student ID
	student_hash = hashlib.sha256(f"{election.id}:{student_id}".encode("utf-8")).hexdigest()

	# Check double-voting (skip for test accounts)
	is_test = request.user.is_superuser or getattr(getattr(request.user, 'profile', None), 'exclude_votes', False)
	if not is_test:
		if VoterRegistration.objects.filter(election=election, student_id_hash=student_hash).exists():
			return JsonResponse({"detail": "This student has already voted in this election."}, status=400)

	# Cancel any existing pending sessions on this kiosk
	KioskSession.objects.filter(
		kiosk_id=kiosk_id,
		status__in=[KioskSession.STATUS_PENDING, KioskSession.STATUS_ACTIVE],
	).update(status=KioskSession.STATUS_CANCELLED)

	# Create new kiosk session (dynamic window based on election settings)
	session = KioskSession.objects.create(
		election=election,
		kiosk_id=kiosk_id,
		student_id_hash=student_hash,
		activated_by=request.user,
		expires_at=timezone.now() + timezone.timedelta(seconds=election.kiosk_timeout),
	)

	return JsonResponse({
		"status": "activated",
		"session_id": session.id,
		"kiosk_id": kiosk_id,
		"expires_at": session.expires_at.isoformat(),
	})


@login_required
@require_POST
def api_kiosk_session_complete(request):
	"""Called by kiosk after ballot is submitted or cancelled/expired — marks KioskSession status."""
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	try:
		payload = json.loads(request.body.decode("utf-8"))
	except Exception:
		return JsonResponse({"detail": "Invalid JSON."}, status=400)

	session_id = payload.get("session_id")
	ballot_id = payload.get("ballot_id")
	status_param = payload.get("status", KioskSession.STATUS_DONE)

	if not session_id:
		return JsonResponse({"detail": "session_id required."}, status=400)

	try:
		session = KioskSession.objects.get(id=session_id, status=KioskSession.STATUS_ACTIVE)
	except KioskSession.DoesNotExist:
		return JsonResponse({"detail": "Session not found or already completed."}, status=404)

	if status_param not in [KioskSession.STATUS_DONE, KioskSession.STATUS_EXPIRED, KioskSession.STATUS_CANCELLED]:
		status_param = KioskSession.STATUS_DONE

	session.status = status_param
	if ballot_id and status_param == KioskSession.STATUS_DONE:
		try:
			session.ballot = Ballot.objects.get(id=ballot_id)
		except Ballot.DoesNotExist:
			pass
	session.save()

	# Register voter (if not test session and status is done)
	if status_param == KioskSession.STATUS_DONE:
		is_test = session.activated_by and (session.activated_by.is_superuser or getattr(getattr(session.activated_by, 'profile', None), 'exclude_votes', False))
		if not is_test:
			VoterRegistration.objects.get_or_create(
				election=session.election,
				student_id_hash=session.student_id_hash,
			)

	return JsonResponse({"status": session.status})


@login_required
def invigilator_dashboard(request):
	"""Invigilator-facing dashboard to manage voter sessions and activate kiosks."""
	profile = getattr(request.user, 'profile', None)
	if not (request.user.is_superuser or (profile and not profile.is_locked)):
		return redirect('login')

	if profile and profile.role == 'kiosk':
		if profile.school_slug:
			return redirect("kiosk-school", school_slug=profile.school_slug)
		return redirect("kiosk")

	school_slug = getattr(profile, 'school_slug', '') if profile else ''
	if request.user.is_superuser:
		school_slug = school_slug or None
	election = _active_election_or_none(school_slug=school_slug or None)

	# Active kiosks (pinged in last 15s)
	active_threshold = timezone.now() - timezone.timedelta(seconds=15)
	if election:
		active_kiosks_list = KioskPresence.objects.filter(
			election=election,
			last_seen_at__gte=active_threshold,
			user__is_superuser=False,
		).select_related("user", "user__profile")
		active_kiosks = [k for k in active_kiosks_list if not (getattr(k.user, 'profile', None) and k.user.profile.exclude_votes)]
		voters_count = VoterRegistration.objects.filter(election=election).count()
		total_students = Student.objects.filter(election=election).count()
	else:
		active_kiosks = []
		voters_count = 0
		total_students = 0

	return render(request, "voting/invigilator.html", {
		"election": election,
		"active_kiosks": active_kiosks,
		"voters_count": voters_count,
		"total_students": total_students,
	})


@login_required
@require_GET
def api_invigilator_kiosks(request):
	"""
	Returns the active kiosks for the invigilator dashboard.
	Available to authenticated invigilators/teachers.
	"""
	profile = getattr(request.user, 'profile', None)
	if not (request.user.is_superuser or (profile and not profile.is_locked and profile.role in ('invigilator', 'teacher'))):
		return JsonResponse({"detail": "Access denied."}, status=403)

	school_slug = getattr(profile, 'school_slug', '') if profile else ''
	if request.user.is_superuser:
		school_slug = request.GET.get("school_slug", "").strip() or school_slug or None

	election = _active_election_or_none(school_slug=school_slug or None)
	
	active_threshold = timezone.now() - timezone.timedelta(seconds=15)
	if election:
		presences_list = KioskPresence.objects.filter(
			election=election,
			last_seen_at__gte=active_threshold,
			user__is_superuser=False,
		).select_related("user", "user__profile").order_by("-last_seen_at")
		presences = [p for p in presences_list if not (getattr(p.user, 'profile', None) and p.user.profile.exclude_votes)]
	else:
		presences = []

	data = []
	for p in presences:
		active_session = KioskSession.objects.filter(
			election=election,
			kiosk_id=p.user.username,
			status__in=[KioskSession.STATUS_PENDING, KioskSession.STATUS_ACTIVE],
			expires_at__gt=timezone.now()
		).first()

		session_data = None
		if active_session:
			student_details = _student_for_hash(election, active_session.student_id_hash)
			session_data = {
				"session_id": active_session.id,
				"status": active_session.status,
				"expires_in_seconds": max(0, int((active_session.expires_at - timezone.now()).total_seconds())),
				"student": student_details or {
					"name": "Unknown Student",
					"student_class": "-",
					"division": "-",
					"student_id": "-"
				}
			}

		data.append({
			"username": p.user.username,
			"kiosk_id": p.user.username,
			"last_seen_seconds_ago": int((timezone.now() - p.last_seen_at).total_seconds()),
			"active_session": session_data,
		})
	return JsonResponse({"active_kiosks": data})


@login_required
@require_GET
def api_invigilator_students(request):
	"""
	Returns the student roster with voting status for the invigilator dashboard.
	Available to authenticated invigilators/teachers (not just superusers).
	"""
	profile = getattr(request.user, 'profile', None)
	if not (request.user.is_superuser or (profile and not profile.is_locked and profile.role in ('invigilator', 'teacher'))):
		return JsonResponse({"detail": "Access denied."}, status=403)

	school_slug = request.GET.get("school_slug", "").strip() or None
	if not request.user.is_superuser and profile and profile.school_slug:
		school_slug = profile.school_slug

	election = _active_election_or_none(school_slug=school_slug)
	if not election:
		return JsonResponse({"students": [], "election": None})

	students = Student.objects.filter(election=election)

	# Get voted hashes
	import hashlib
	voted_hashes = set(
		VoterRegistration.objects.filter(election=election).values_list("student_id_hash", flat=True)
	)

	data = []
	for s in students:
		val = s.student_id.lower() if s.student_id else f"student_id__{s.id}"
		h = hashlib.sha256(f"{election.id}:{val}".encode()).hexdigest()
		has_voted = h in voted_hashes
		data.append({
			"id": s.id,
			"name": s.name,
			"student_class": s.student_class,
			"division": s.division,
			"student_id": s.student_id or "",
			"has_voted": has_voted,
		})

	return JsonResponse({
		"students": data,
		"election_id": election.id,
		"school_slug": election.school_slug,
		"total": len(data),
		"voted": sum(1 for d in data if d["has_voted"]),
	})


@login_required
@require_POST
def api_invigilator_student_create(request):
	"""
	Allows teachers to add a student to their active election roster.
	"""
	profile = getattr(request.user, 'profile', None)
	if not request.user.is_superuser:
		if not profile or profile.is_locked or profile.role != 'teacher':
			return JsonResponse({"detail": "Access denied. Only teacher accounts can add students."}, status=403)

	school_slug = getattr(profile, 'school_slug', '') if profile else ''
	if request.user.is_superuser:
		school_slug = request.POST.get("school_slug", "").strip() or school_slug or None

	election = _active_election_or_none(school_slug=school_slug or None)
	if not election:
		return JsonResponse({"detail": "No active election found for your school."}, status=400)

	name = request.POST.get("name", "").strip()
	student_class = request.POST.get("student_class", "").strip()
	division = request.POST.get("division", "").strip()
	student_id = request.POST.get("student_id", "").strip()

	if not name:
		return JsonResponse({"detail": "Student name is required."}, status=400)
	if not student_class:
		return JsonResponse({"detail": "Class is required."}, status=400)

	if student_id and Student.objects.filter(election=election, student_id__iexact=student_id).exists():
		return JsonResponse({"detail": "A student with this ID already exists in this election."}, status=400)

	student = Student.objects.create(
		election=election,
		name=name,
		student_class=student_class,
		division=division,
		student_id=student_id
	)

	return JsonResponse({
		"status": "success",
		"student": {
			"id": student.id,
			"name": student.name,
			"student_class": student.student_class,
			"division": student.division,
			"student_id": student.student_id,
			"has_voted": False
		}
	})


@login_required
@require_POST
def api_invigilator_students_import(request):
	"""
	Allows teachers to bulk import students into their active election roster.
	Superusers may pass school_slug to target a specific active school election.
	"""
	election, error_response = _teacher_student_election(request)
	if error_response:
		return error_response

	try:
		result = _import_students_from_upload(election, request.FILES.get("excel_file"))
	except Exception as exc:
		return JsonResponse({"detail": str(exc)}, status=400)

	return JsonResponse({
		"status": "success",
		"created": result["created"],
		"skipped": result["skipped"],
		"errors": result["errors"],
	})


@login_required
@require_GET
def api_invigilator_students_import_template(request):
	"""Download a blank student import template for teacher uploads."""
	election, error_response = _teacher_student_election(request)
	if error_response:
		return error_response

	file_format = request.GET.get("format", "xlsx").lower()
	if file_format not in ("xlsx", "csv"):
		return JsonResponse({"detail": "Unsupported template format."}, status=400)
	return _student_import_template_response(election, file_format)


@login_required
@require_GET
def api_session_status(request, session_id):
	device_error = _check_kiosk_device(request)
	if device_error:
		return device_error

	try:
		session = KioskSession.objects.get(id=session_id)
		return JsonResponse({"status": session.status})
	except KioskSession.DoesNotExist:
		return JsonResponse({"detail": "Session not found."}, status=404)


@login_required
@require_POST
def api_invigilator_cancel_session(request):
	"""
	Invigilator explicitly cancels an active/pending kiosk session.
	Called when the invigilator closes the activation modal or the countdown expires.
	The kiosk's health-check poll will detect the cancellation and return to the waiting screen.
	"""
	profile = getattr(request.user, 'profile', None)
	if not request.user.is_superuser:
		if not profile or profile.is_locked or profile.role not in ('invigilator', 'teacher'):
			return JsonResponse({"detail": "Access denied."}, status=403)

	try:
		payload = json.loads(request.body.decode("utf-8"))
	except Exception:
		return JsonResponse({"detail": "Invalid JSON."}, status=400)

	session_id = payload.get("session_id")
	if not session_id:
		return JsonResponse({"detail": "session_id required."}, status=400)

	updated = KioskSession.objects.filter(
		id=session_id,
		status__in=[KioskSession.STATUS_PENDING, KioskSession.STATUS_ACTIVE],
	).update(status=KioskSession.STATUS_CANCELLED)

	if updated:
		return JsonResponse({"status": "cancelled"})
	# Session already completed or does not exist — treat as success (idempotent)
	return JsonResponse({"status": "already_resolved"})
